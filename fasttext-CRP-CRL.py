#!/usr/bin/env python
# coding: utf-8

# In[170]:


import stem_module as sm
from transformers import AutoModel, AutoTokenizer, BertTokenizer
import random
import numpy as np
from itertools import combinations
import torch
from gensim.models import KeyedVectors
from gensim.scripts.glove2word2vec import glove2word2vec 
from sklearn.metrics.pairwise import cosine_similarity as cs


# # FastText

# Completely unrelated words would have cos θ ≈ 0, and strong associates would have cos θ values between 0.4 and 1.0.

# In[171]:


ft= sm.loadModel(r'cc.tr.300.bin') 


# In[172]:


word_list= sm.getWordList("kelime_listesi_burak.txt") 


# In[173]:


vectors = sm.get_word_vectors(word_list, ft)


# In[174]:


word_distances = sm.cos_similarity_function(vectors)


# In[175]:


#dissimilars_ft= sm.generate_one_list(ft,word_distances,word_list)
dissimilars_ft=[]
i=0
while i<10:
    dissimilars_ft.append(sm.generate_one_list(ft,word_distances,word_list,12))
    i=i+1
print(dissimilars_ft)


# # Word2vec

# In[176]:


word2vec_model = KeyedVectors.load_word2vec_format('trmodelword2vec', binary=True)


# #### word2vec benzemez

# In[177]:


res_word2vec=[]
word2vec_word_list=[]
for i in word_list:
    if i in word2vec_model:
        res_word2vec.append((i,word2vec_model[i]))
        word2vec_word_list.append(i)
        
    else:
        continue
word2vec_word_list= np.array(word2vec_word_list)    


# In[178]:


res_word2vec


# In[179]:


len(res_word2vec)


# In[180]:


word2vec_vectors= dict(res_word2vec)


# In[181]:


word_distances1 = sm.cos_similarity_function(np.array(list(word2vec_vectors.values())))


# In[182]:


word_distances1


# In[183]:


dissimilar_word2vec=[]
i=0
while i<10:
    dissimilar_word2vec.append(sm.generate_one_list(word2vec_model,word_distances1,word2vec_word_list))
    i=i+1
print(dissimilar_word2vec)


# #### benzer listeler

# In[184]:


similars_word2vec= sm.generate_similar_words(word2vec_word_list,word_distances1,12,print_values=False)
similars_word2vec


# In[185]:


similars_fastText= sm.generate_similar_words(word_list,word_distances,12,print_values=False)
similars_fastText


# ### Diagnosis For  similars

# In[186]:


for a,b in combinations(similars_word2vec[0],2):
  print((a,b,sm.cosine_similarity(word2vec_model[a],word2vec_model[b])))


# In[187]:


def show_diagnosis(similars_word2vec,word2vec_model,show_cos_values=True):
    similarities = []
    for index in range(len(similars_word2vec)):
        current_similarity = []  
        for a, b in combinations(similars_word2vec[index], 2):
            similarity = sm.cosine_similarity(word2vec_model[a], word2vec_model[b])
            if(show_cos_values==True):
                current_similarity.append((a, b, similarity))
            else:
                 current_similarity.append((a, b))
        similarities.append(current_similarity)
    return similarities


# In[188]:


diag_similar_w2v=  show_diagnosis(similars_word2vec,word2vec_model)


# In[189]:


diag_similar_w2v1=  show_diagnosis(similars_word2vec,word2vec_model,show_cos_values=False)
                                  


# In[190]:


diag_dissimilar_w2v= show_diagnosis(dissimilar_word2vec,word2vec_model)


# In[191]:


diag_dissimilar_w2v1= show_diagnosis(dissimilar_word2vec,word2vec_model,show_cos_values=False)
                                   


# In[192]:


dissimilar_word2vec


# In[197]:


diag_similar_ft= show_diagnosis(similars_fastText,ft)


# In[198]:


diag_similar_ft1= show_diagnosis(similars_fastText,ft,show_cos_values=False)
                                   


# In[199]:


diag_dissimilar_ft= show_diagnosis(dissimilars_ft,ft)


# In[200]:


diag_dissimilar_ft1= show_diagnosis(dissimilars_ft,ft,show_cos_values=False)
                                   


# In[208]:


for eleman in diag_similar_w2v1[:]:  
    for i in range(55):
        print(eleman[i])


# In[209]:


for eleman in diag_similar_w2v1[:]:  
    for i in range(len(eleman)):
        print(eleman[i])


# ### Save Results

# In[201]:


def save_results(filename,liste):
    with open(filename, 'w') as dosya:
        for eleman in liste[:]:  
            for i in range(55):
                dosya.write(str(eleman[i]) + "\n")


# In[202]:


save_results("similar_word2vec_list.txt",diag_similar_w2v1)


# In[203]:


save_results("similar_ft_list.txt",diag_similar_ft1)


# In[204]:


save_results("dissimilar_word2vec_list.txt",diag_dissimilar_w2v1)


# In[205]:


save_results("dissimilar_ft_list.txt",diag_dissimilar_ft1)


# ## DENEYE HAZIRLIK

# In[7]:


import openpyxl


# In[34]:


for index, sublist in enumerate(res, start=1):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["test_words"])

    for word in sublist:
        sheet.append([word])

    filename = f'study_word_{index}.xlsx'
    workbook.save(filename)


# In[142]:


def random_rakamlar_virgulle():
    sonuc = ', '.join([str(random.randint(0, 9)) + "+" + str(random.randint(0, 9)) for _ in range(100)])
    random_rakamlar_liste = sonuc.split(', ')
    return random_rakamlar_liste

random_rakamlar = random_rakamlar_virgulle()


# In[146]:


# Yeni bir Excel dosyası oluşturun
workbook1 = openpyxl.Workbook()

# Varsayılan çalışma sayfasını seçin
sheet1 = workbook1.active

sheet1.append(random_rakamlar)

workbook1.save('arithmatic_test.xlsx')


# In[163]:


def random_rakamlar_virgulle():
    sonuc=[]
    sonuc = ', '.join([str(random.randint(0, 9)) + "+" + str(random.randint(0, 9)) for _ in range(100)])
    random_rakamlar_liste = sonuc.split(', ')
    return random_rakamlar_liste

random_rakamlar = random_rakamlar_virgulle()

# Yeni bir Excel dosyası oluşturun
workbook1 = openpyxl.Workbook()

# Varsayılan çalışma sayfasını seçin
sheet1 = workbook1.active

# Verileri Excel çalışma sayfasına yazın
for item in random_rakamlar:
    sheet1.append([item])

# Excel dosyasını kaydedin
workbook1.save('arithmatic_test.xlsx')


# In[164]:


data = pd.read_excel('arithmatic_test.xlsx')
df["summation"] = data


# In[166]:


file_path = 'arithmatic_test.xlsx'

# Save the DataFrame to an Excel file
df.to_excel(file_path, index=False)  # Set index=False to omit the index column

print(f'DataFrame saved to {file_path}')


# In[24]:


def alfabe_harfleri_virgulle():
    alfabe = "abcdefghijklmnopqrstuvwxyz"
    sonuc = ','.join(["'" + harf + "'" for harf in alfabe])
    return sonuc
alfabe_harfleri = alfabe_harfleri_virgulle()
print(alfabe_harfleri)


# In[9]:


sm.generate_similar_words(word_list,word_distances,12)


# # fastText-CRP (semantic-CRP)

# for instance we have a study list with 12 item but the subject recall some of them (but not all)
# 
# - calculate cos_u from each successfully recalled words and rank them.
# - we first need to discretize the cosine similarity values into 100 bins of equal width. We can do this by dividing the range of cosine similarity values into 100 intervals of width. 
# - subjects are more likely to make transitions to words with a higher value of cos u relative to the just-recalled word.
# - we take actual and possible counts to find CRP (Actual/Possible Count) [https://memory.psych.upenn.edu/CRP_Tutorial]
# 
# Note: it is a single-trial example
# 

# In[10]:


# for a list taken randomly from created before:
# I print random number to find possible transitions.
items_recalled=[]
rand_num=random.randint(0, 12)
print(rand_num)
rand_list= res[rand_num]

for i in range(8):
    rand_num=random.randint(0, 11)
    if i in items_recalled:
        continue
    elif(len(items_recalled)==8):
        break
    else:
        items_recalled.append(rand_list[rand_num])

items_recalled= set(items_recalled)


# In[26]:


items_recalled= list(items_recalled)


# In[27]:


res[2]


# In[28]:


items_recalled #1st to last recalled order


# In[29]:


diag_actual= sm.show_diagnosis(ft,items_recalled,True) # successively recalled words.
diag_actual


# In[30]:


diag_possible= sm.show_diagnosis(ft,res[9],True)
diag_possible #possible transitions


# In[31]:


len(diag_actual)


# In[32]:


# dividing the range of cosine similarity values  into 100 intervals
def define_bins(diag_):
    first= diag_[0][2]
    last= diag_[-1][2]
    width= (last - first)/100
    # Create a dictionary to store the values to bins
    bin_discrete = {}
    counter=0
    # Use a for loop to create 100 variables and assign values
    for i in range(1, 100):
        bin_discrete[f'Bin{i}'] = (first+width*i,first+width*(i+1),counter)
    bin_discrete["Bin0"]= (first,first+width,counter)
    return bin_discrete


# In[33]:


bin_discrete= define_bins(diag_actual)
bin_discrete 


# In[34]:


#compare and assign values to bins
# Value to check
def compare_and_assign_value(value_to_check,bin_discrete,show_results= False):
    # Iterate through the bins and check if the value is within the range
    in_bin = None  # Initialize a variable to store the bin where the value is found

    for bin_name, (min_value, max_value,counter) in bin_discrete.items():
        if min_value <= value_to_check <= max_value:
            in_bin = bin_name
            bin_discrete[bin_name] = (min_value, max_value, counter + 1)
            break  # Exit the loop when the value is found in a bin

    # Check the result
    if in_bin:
        print(f"The value {value_to_check} is in {in_bin} .") #and count: {counter}
    else:
        print(f"The value {value_to_check} is not in any bin.")
    return (value_to_check,in_bin)


# In[37]:


#actual transitions
result=[]
for i in range(len(diag_actual)):
    result.append(compare_and_assign_value(diag_actual[i][2],bin_discrete,show_results= False))


# In[38]:


result #actual transitions results


# # fastText-CRL (semantic-CRL)

# - The conditional response latency as a function of mean LSA cos u for each bin
# - inspects shows the mean inter-response time (IRT) between successive recalls of the bin of the cos u distribution
# - LSA cos u also affects IRTs (inter-response-times) in free recall. (so does fasttext & BERT)
# - IRTs are shorter when the successively recalled words are similar (i.e., have high cos u).
# - subjects are more likely to recall items that are semantically related to the just-recalled item
# 
# Note: it is a single-trial example

# In[23]:


items_recalled


# In[42]:


diag_actual


# In[44]:


bin_discrete


# In[45]:


result


# In[76]:


len(result)


# In[99]:


items=[]
for i,y in result:
    items.append(i)
items


# In[96]:


time=  np.random.randint(1, 11, len(result))
mean_IRTs = np.array(time)
mean_IRTs= np.sort(mean_IRTs)[::-1]
mean_IRTs


# In[74]:


mean_IRTs


# In[75]:


x


# In[100]:


import matplotlib.pyplot as plt

# Generate random data for the x-axis (e.g., 8 data points)


# Create a dot plot
plt.scatter(items, mean_IRTs, marker='o', color='b')

plt.ylim(0, 10.5)
plt.xlim(bin_discrete['Bin0'][0], bin_discrete['Bin99'][1])


# Add labels and title
plt.xlabel('Cosine')
plt.ylabel('Conditional Response Latency')
plt.title('Semantic-CRL')

# Show the plot
plt.show()


# # BERT MODELİ

# Turkish pre-trained model: https://huggingface.co/dbmdz/bert-base-turkish-cased

# In[14]:


bert_model = AutoModel.from_pretrained("dbmdz/bert-base-turkish-cased")


# In[101]:


vector1= sm.calculate_bert_embeddings_model(word_list)


# In[102]:


vector1


# In[103]:


word_distances1 = sm.cos_similarity_function(vector1)


# In[104]:


sm.generate_similar_words(word_list,word_distances1,12)

